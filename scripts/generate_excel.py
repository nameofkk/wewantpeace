"""기자 발송 리스트 + 템플릿을 탭 분리 엑셀로 생성"""
import csv
import html
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT = "/mnt/c/Users/krshi/OneDrive/바탕 화면/기자_콜드이메일_발송키트.xlsx"
CSV_PATH = "docs/marketing/cold-email-send-list.csv"

wb = Workbook()

# ── 스타일 ──
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
label_font = Font(bold=True, size=12, color="2F5496")
body_font = Font(size=11)

def style_header(ws, cols):
    for col_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

# ═══════════════════════════════════════
# 탭 1: 발송 리스트
# ═══════════════════════════════════════
ws_list = wb.active
ws_list.title = "발송리스트"

cols = ["순번", "매체", "이름", "이메일", "추천템플릿", "제목", "대표기사(참고용)", "비고"]
style_header(ws_list, cols)

with open(CSV_PATH, encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for i, row in enumerate(reader, 2):
        ws_list.cell(row=i, column=1, value=int(row["순번"])).border = thin_border
        ws_list.cell(row=i, column=2, value=row["매체"]).border = thin_border
        ws_list.cell(row=i, column=3, value=row["이름"]).border = thin_border
        ws_list.cell(row=i, column=4, value=row["이메일"]).border = thin_border
        ws_list.cell(row=i, column=5, value=row["추천템플릿"]).border = thin_border
        ws_list.cell(row=i, column=6, value=row["제목"]).border = thin_border
        ws_list.cell(row=i, column=7, value=row["대표기사"]).border = thin_border
        ws_list.cell(row=i, column=8, value=row["비고"]).border = thin_border

ws_list.column_dimensions["A"].width = 6
ws_list.column_dimensions["B"].width = 12
ws_list.column_dimensions["C"].width = 10
ws_list.column_dimensions["D"].width = 28
ws_list.column_dimensions["E"].width = 12
ws_list.column_dimensions["F"].width = 45
ws_list.column_dimensions["G"].width = 55
ws_list.column_dimensions["H"].width = 25

# ═══════════════════════════════════════
# 템플릿 시트 생성 함수
# ═══════════════════════════════════════
def create_template_sheet(wb, tab_name, target, subject, body, followup_subject, followup_body):
    ws = wb.create_sheet(title=tab_name)

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value=f"템플릿 {tab_name}").font = Font(bold=True, size=14, color="2F5496")
    r += 1
    ws.cell(row=r, column=1, value="대상:").font = label_font
    ws.cell(row=r, column=2, value=target).font = body_font
    r += 2

    # 제목
    ws.cell(row=r, column=1, value="이메일 제목").font = label_font
    r += 1
    ws.cell(row=r, column=1, value=subject).font = Font(size=11, bold=True)
    r += 2

    # 본문
    ws.cell(row=r, column=1, value="이메일 본문 (복사용)").font = label_font
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell = ws.cell(row=r, column=1, value=body)
    cell.font = body_font
    cell.alignment = wrap
    ws.row_dimensions[r].height = 280
    r += 2

    # 치환 안내
    ws.cell(row=r, column=1, value="치환 필수 항목").font = label_font
    r += 1
    ws.cell(row=r, column=1, value="{이름}").font = Font(size=11, bold=True, color="C00000")
    ws.cell(row=r, column=2, value="기자 이름 (예: 홍길동)").font = body_font
    r += 1
    ws.cell(row=r, column=1, value="{최근기사제목}").font = Font(size=11, bold=True, color="C00000")
    ws.cell(row=r, column=2, value="해당 기자가 최근 쓴 기사 제목 — 반드시 실제로 읽고 채울 것").font = body_font
    r += 1
    ws.cell(row=r, column=1, value="{코멘트}").font = Font(size=11, bold=True, color="C00000")
    ws.cell(row=r, column=2, value="기사에 대한 구체적 감상 1문장 (예: 호르무즈 파병 관련 각국 입장 정리가 인상적이었습니다)").font = body_font
    r += 2

    # 팔로업
    ws.cell(row=r, column=1, value="팔로업 이메일 (3영업일 후, 1회만)").font = label_font
    r += 1
    ws.cell(row=r, column=1, value="제목:").font = Font(size=11, bold=True)
    ws.cell(row=r, column=2, value=followup_subject).font = body_font
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell = ws.cell(row=r, column=1, value=followup_body)
    cell.font = body_font
    cell.alignment = wrap
    ws.row_dimensions[r].height = 120

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 30

# ═══════════════════════════════════════
# A안
# ═══════════════════════════════════════
create_template_sheet(
    wb, "A안-국제정치부",
    "연합뉴스, 한겨레, 중앙일보, 조선일보, 동아일보 — 국제부/정치부 기자",
    "195개국 분쟁 실시간 모니터링 — 취재 참고 도구 공유",
    """{이름} 기자님 안녕하세요.

최근 쓰신 "{최근기사제목}" 기사 잘 읽었습니다. {기사에 대한 구체적 코멘트 1문장}

관련해서 취재에 참고가 될 수 있는 도구가 있어서 공유드립니다.

WeWantPeace(https://www.wewantpeace.live)라는 플랫폼인데, 로이터·AP부터 텔레그램 OSINT 채널까지 66개 소스를 5분마다 수집해서 195개국 분쟁 상황을 한 화면에 보여줍니다. 구글 플레이스토어 뉴스 카테고리 1위, 디스콰이엇 이주의 프로덕트 수상 이력이 있습니다.

3월 이란-미국 긴장 고조 때 걸프 현지 채널이 국제 통신사보다 수 시간 먼저 신호를 보낸 걸 이 플랫폼에서 확인했는데, 속보 판단하실 때 교차 확인용으로 쓸 만합니다.

무료이고 가입 없이 바로 확인 가능합니다. 혹시 관련 취재에 관심 있으시면 추가 자료나 인터뷰 드리겠습니다.

김신혁 드림
WeWantPeace 대표
krshin7@naver.com""",
    "Re: 195개국 분쟁 실시간 모니터링 — 취재 참고 도구 공유",
    """{이름} 기자님, 지난번 메일 보셨는지 확인 차 한 번 더 연락드립니다.

WeWantPeace(https://www.wewantpeace.live) — 195개국 분쟁 실시간 모니터링 플랫폼입니다.
가입 없이 바로 확인 가능합니다.

관심 없으시면 무시하셔도 됩니다. 바쁘신 와중에 감사합니다.

김신혁 드림
krshin7@naver.com""",
)

# ═══════════════════════════════════════
# B안
# ═══════════════════════════════════════
create_template_sheet(
    wb, "B안-경제산업부",
    "매일경제, 한국경제, 헤럴드경제, 아시아경제 — 경제부/산업부 기자",
    "분쟁 리스크가 한국 경제에 미치는 영향, 실시간 수치로 봅니다",
    """{이름} 기자님 안녕하세요.

"{최근기사제목}" 기사 읽었습니다. {기사에 대한 구체적 코멘트 1문장}

취재에 참고가 될 수 있는 서비스가 있어서 공유드립니다.

WeWantPeace(https://www.wewantpeace.live)는 66개 국제 소스를 5분 간격으로 수집해서 195개국 분쟁 상황을 실시간으로 보여주는 플랫폼입니다. 플레이스토어 뉴스 앱 1위, 디스콰이엇 이주의 프로덕트를 수상했습니다.

경제 취재에 쓸 만한 기능이 있는데, KScore라는 자체 지표로 각 분쟁이 한국에 미치는 영향도를 0~10점으로 수치화합니다. 호르무즈 해협 이벤트 급증 → 유가 영향, 대만해협 군사 동향 → 반도체 공급망 리스크 같은 흐름을 데이터로 확인할 수 있어서 리스크 기사 쓰실 때 참고가 됩니다.

무료이고 가입 없이 사용 가능합니다. 관심 있으시면 편하게 연락 주세요.

김신혁 드림
WeWantPeace 대표
krshin7@naver.com""",
    "Re: 분쟁 리스크가 한국 경제에 미치는 영향, 실시간 수치로 봅니다",
    """{이름} 기자님, 지난번 메일 보셨는지 확인 차 한 번 더 연락드립니다.

WeWantPeace(https://www.wewantpeace.live) — 195개국 분쟁 실시간 모니터링 플랫폼입니다.
가입 없이 바로 확인 가능합니다.

관심 없으시면 무시하셔도 됩니다. 바쁘신 와중에 감사합니다.

김신혁 드림
krshin7@naver.com""",
)

# ═══════════════════════════════════════
# C안
# ═══════════════════════════════════════
create_template_sheet(
    wb, "C안-IT테크부",
    "전자신문 — IT/테크/스타트업 섹션 기자",
    "플레이스토어 뉴스 1위 앱 — 스타트업 출신 개발자의 분쟁 모니터링 서비스",
    """{이름} 기자님 안녕하세요.

"{최근기사제목}" 기사 인상 깊게 읽었습니다. {기사에 대한 구체적 코멘트 1문장}

기사 소재가 될 수 있을 것 같아 연락드립니다.

WeWantPeace(https://www.wewantpeace.live)는 195개국 분쟁·안보 상황을 실시간으로 수집·분석하는 플랫폼입니다. 로이터·AP부터 텔레그램 OSINT 채널까지 66개 소스를 5분 간격으로 모니터링합니다.

구글 플레이스토어 뉴스 카테고리 인기 앱 1위를 달성했고, 디스콰이엇에서 이주의 프로덕트로 선정됐습니다.

저는 18세에 첫 창업을 시작해서 두 번째 사업에서 연 매출 50억을 만든 경험이 있습니다. 2024년 12월 계엄 사태를 겪으면서 국제 안보 상황을 일반인도 쉽게 파악할 수 있는 도구가 필요하다고 느꼈고, 그 경험을 바탕으로 이 서비스를 만들었습니다.

무료이고 가입 없이 바로 사용할 수 있습니다. 취재에 관심 있으시면 인터뷰나 자료 드리겠습니다.

김신혁 드림
WeWantPeace 대표
krshin7@naver.com""",
    "Re: 플레이스토어 뉴스 1위 앱 — 스타트업 출신 개발자의 분쟁 모니터링 서비스",
    """{이름} 기자님, 지난번 메일 보셨는지 확인 차 한 번 더 연락드립니다.

WeWantPeace(https://www.wewantpeace.live) — 195개국 분쟁 실시간 모니터링 플랫폼입니다.
가입 없이 바로 확인 가능합니다.

관심 없으시면 무시하셔도 됩니다. 바쁘신 와중에 감사합니다.

김신혁 드림
krshin7@naver.com""",
)

# ═══════════════════════════════════════
# 발송 가이드 탭
# ═══════════════════════════════════════
ws_guide = wb.create_sheet(title="발송가이드")
guide_items = [
    ("발송 규칙", ""),
    ("하루 최대", "50통 — 스팸 필터 회피"),
    ("발송 시간", "오전 9~11시 — 기자 이메일 확인 피크"),
    ("팔로업", "3영업일 후 1회만 — 그 이상은 스팸"),
    ("수신거부", "요청 시 즉시 처리"),
    ("", ""),
    ("절대 하지 말 것", ""),
    ("과장 표현 금지", '"혁신적", "세계 최초", "획기적" → 기자 신뢰 영구 파괴'),
    ("이미지 첨부 금지", "첫 메일에 이미지/파일 첨부 → 스팸 필터 + 기자가 안 열어봄"),
    ("대량 발송 티", '{최근기사제목}을 실제로 안 채우면 → "기사 잘 읽고 있습니다" 수준의 뻔한 메일'),
    ("", ""),
    ("발송 전 체크리스트", ""),
    ("1", "해당 기자의 최근 기사 1건을 실제로 읽었는가?"),
    ("2", "{최근기사제목}과 {코멘트}를 실제 내용으로 채웠는가?"),
    ("3", "{이름}을 정확히 치환했는가?"),
    ("4", "해당 기자의 분야와 템플릿(A/B/C)이 맞는가?"),
]

for i, (col1, col2) in enumerate(guide_items, 1):
    c1 = ws_guide.cell(row=i, column=1, value=col1)
    c2 = ws_guide.cell(row=i, column=2, value=col2)
    if col2 == "" and col1:
        c1.font = Font(bold=True, size=13, color="2F5496")
    else:
        c1.font = Font(bold=True, size=11)
        c2.font = body_font

ws_guide.column_dimensions["A"].width = 22
ws_guide.column_dimensions["B"].width = 70

wb.save(OUTPUT)
print(f"엑셀 생성 완료: {OUTPUT}")
print(f"탭: 발송리스트, A안-국제정치부, B안-경제산업부, C안-IT테크부, 발송가이드")
